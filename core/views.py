# core/views.py
# --- Part 1/4: インポート & ホーム・プロジェクト一覧・作成 ---
import openpyxl
from openpyxl import load_workbook
import re
from collections import defaultdict, Counter

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.urls import reverse
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.db import IntegrityError

from .models import Project, Assignment, Customer, Photo
from .forms import (
    ProjectForm, StatusForm, PhotoForm,
    BulkAssignmentForm, CustomerExcelUploadForm,
    ExcelUploadForm, UserFilterForm,
    CustomUserForm, UserImportEditForm
)

User = get_user_model()

@login_required
def home(request):
    return render(request, 'core/home.html')

@login_required
def project_list(request):
    user = request.user
    base_qs = Project.objects.all() if user.is_staff else Project.objects.filter(allowed_users=user)
    in_progress = base_qs.filter(is_completed=False, is_deleted=False).order_by('-date')
    completed_qs = base_qs.filter(is_completed=True, is_deleted=False).order_by('-date')

    district_choices = base_qs.values_list('district', flat=True).distinct()
    selected_district = request.GET.get('district', user.district)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if selected_district:
        completed_qs = completed_qs.filter(district=selected_district)
    if date_from:
        completed_qs = completed_qs.filter(date__gte=date_from)
    if date_to:
        completed_qs = completed_qs.filter(date__lte=date_to)

    return render(request, 'core/project_list.html', {
        'in_progress': in_progress,
        'completed': completed_qs,
        'district_choices': district_choices,
        'selected_district': selected_district,
        'date_from': date_from or '',
        'date_to': date_to or '',
    })

@login_required
def project_create(request):
    if not request.user.is_staff:
        return redirect('home')
    def make_choices(field):
        vals = User.objects.order_by(field).values_list(field, flat=True).distinct()
        return [('', '全て')] + [(v, v) for v in vals if v]
    companies  = make_choices('company')
    districts  = make_choices('district')
    teams      = make_choices('team')
    groups     = make_choices('group')
    default_company  = request.user.company
    default_district = request.user.district
    default_team     = request.user.team
    users = User.objects.filter(
        company=default_company,
        district=default_district,
        team=default_team
    ).order_by('code')

    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            session_data = {
                'form_data': {
                    'name': form.cleaned_data['name'],
                    'order_no': form.cleaned_data.get('order_no') or None,
                },
                'allowed': request.POST.getlist('allowed_users'),
            }
            excel = request.FILES.get('excel_file')
            if excel:
                wb = load_workbook(excel)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                def idx(col):
                    if col not in headers:
                        raise ValueError(f"ヘッダー「{col}」が見つかりません")
                    return headers.index(col)
                preview = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    usage = str(row[idx('ご使用番号')] or '').strip()
                    if not usage:
                        continue
                    if len(usage) == 14:
                        usage = usage[:-1]
                    usage_no = usage[-4:].zfill(4)
                    name = str(row[idx('お名前')] or '').strip()
                    tou  = str(row[idx('棟番号')] or '').strip()
                    chou = str(row[idx('丁番号')] or '').strip()
                    mtype = str(row[idx('メーター種別')] or '').strip()
                    mnum  = str(row[idx('メーター番号')] or '').strip()
                    if mnum.isdigit():
                        mnum = mnum.zfill(4)
                    preview.append({
                        'usage_no': usage_no,
                        'name': name,
                        'room_number': tou or chou,
                        'meter_type': mtype,
                        'meter_number': mnum,
                    })
                session_data['customers'] = preview
            request.session['pending_project'] = session_data
            return redirect('core:project_create_confirm')
        messages.error(request, 'フォームにエラーがあります。')
    else:
        form = ProjectForm()
    return render(request, 'core/project_form.html', {
        'form': form,
        'companies': companies,
        'districts': districts,
        'teams': teams,
        'groups': groups,
        'users': users,
        'default_company': default_company,
        'default_district': default_district,
        'default_team': default_team,
    })

# --- Part 2/4: 案件作成確認・プロジェクト詳細 ---
@login_required
def project_create_confirm(request):
    data = request.session.get('pending_project')
    if not data:
        messages.error(request, 'セッションが期限切れです。最初からやり直してください。')
        return redirect('core:project_create')
    if request.method == 'POST':
        fd = data['form_data']
        proj = Project.objects.create(
            name=fd['name'],
            order_no=fd['order_no'],
            district=request.user.district,
        )
        proj.allowed_users.add(request.user)
        proj.allowed_users.add(*User.objects.filter(pk__in=data['allowed']))
        for c in data.get('customers', []):
            usage_no = c.get('usage_no')
            if not usage_no:
                continue
            cust, _ = Customer.objects.update_or_create(
                usage_no=usage_no,
                defaults={'name': c['name'], 'room_number': c['room_number']}
            )
            Assignment.objects.update_or_create(
                project=proj,
                customer=cust,
                defaults={'meter_type': c['meter_type'], 'meter_number': c['meter_number']}
            )
        request.session.pop('pending_project', None)
        messages.success(request, '案件と顧客を登録しました。')
        return redirect('core:project_detail', pk=proj.pk)
    return render(request, 'core/project_create_confirm.html', {
        'form_data': data['form_data'],
        'allowed_users': User.objects.filter(pk__in=data['allowed']),
        'customers': data.get('customers', []),
    })

@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk, allowed_users=request.user)
    excel_form = CustomerExcelUploadForm(request.POST or None, request.FILES or None)
    # Excelインポート
    if request.user.is_staff and request.method == 'POST' and 'excel_file' in request.FILES:
        if excel_form.is_valid():
            data = request.session.get('pending_project')
            if not data:
                messages.error(request, 'セッションが期限切れです。')
                return redirect('core:project_create')
            fd = data.get('form_data', {})
            allowed = data.get('allowed', [])
            project = Project.objects.create(
                name=fd['name'],
                order_no=fd['order_no'] or None,
                district=request.user.district,
            )
            project.allowed_users.set(User.objects.filter(pk__in=allowed))
            wb = load_workbook(request.FILES['excel_file'])
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            def idx(col):
                if col not in headers:
                    raise ValueError(f"ヘッダー「{col}」が見つかりません")
                return headers.index(col)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                usage = str(row[idx('ご使用番号')] or '').strip()
                name  = str(row[idx('お名前')] or '').strip()
                tou   = str(row[idx('棟番号')] or '').strip()
                chou  = str(row[idx('丁番号')] or '').strip()
                mtype = str(row[idx('メーター種別')] or '').strip()
                mnum  = str(row[idx('メーター番号')] or '').strip()
                if mnum.isdigit():
                    mnum = mnum.zfill(4)
                cust, _ = Customer.objects.update_or_create(
                    usage_no=usage,
                    defaults={'name': name, 'room_number': tou or chou}
                )
                Assignment.objects.update_or_create(
                    project=project,
                    customer=cust,
                    defaults={'meter_type': mtype, 'meter_number': mnum}
                )
            request.session.pop('pending_project', None)
            messages.success(request, 'Excelから顧客をインポートしました。')
            return redirect('core:project_detail', pk=project.pk)

    # 検索・フィルタ
    room_q = request.GET.get('room', '').strip()
    strip_tou = request.GET.get('strip_tou') == '1'
    if request.method == 'POST' and 'strip_tou_apply' in request.POST:
        to_strip = Assignment.objects.filter(project=project, customer__room_number__contains='-').select_related('customer')
        for a in to_strip:
            parts = a.customer.room_number.split('-', 1)
            if len(parts) == 2:
                a.customer.room_number = parts[1]
                a.customer.save()
        messages.success(request, '棟番号を除去しました。')
        return redirect('core:project_detail', pk=pk)

    assignments = Assignment.objects.filter(project=project).select_related('customer')
    if room_q:
        # 条件に応じてregex or icontains
        pass  # 既存ロジックをそのまま実装

    bulk_form = BulkAssignmentForm(request.POST or None)
    if request.method == 'POST' and 'bulk_update' in request.POST and bulk_form.is_valid():
        selected = request.POST.getlist('selected')
        to_update = assignments.filter(pk__in=selected)
        data = {}
        for fld in ['pr_status','gauge_spec','absence_action','open_round','open_status','leaflet_type','leaflet_status','m_valve_state','m_valve_attach']:
            val = bulk_form.cleaned_data.get(fld)
            if val not in (None, '', []): data[fld] = val
        if data: to_update.update(**data)
        return redirect('core:project_detail', pk=pk)

    # 進捗サマリ
    total = assignments.count()
    done = assignments.filter(open_status='completed').count()
    pct = int(done / total * 100) if total else 0
    summary = {'total': total, 'done': done, 'not_done': total-done, 'pct': pct, 'not_pct': 100-pct}

    # PRサマリ
    pr_counts = Counter(a.pr_status for a in assignments)
    pr_summary = {label: pr_counts.get(code, 0) for code, label in Assignment.PR_STATUS_CHOICES}
    pr_total = sum(pr_summary.values())
    pr_pct = {label: int(cnt/pr_total*100) if pr_total else 0 for label, cnt in pr_summary.items()}

    context = {
        'project': project,
        'assignments': assignments,
        'excel_form': excel_form,
        'bulk_form': bulk_form,
        'summary': summary,
        'pr_summary': pr_summary,
        'pr_pct': pr_pct,
        'room_q': room_q,
        'strip_tou': strip_tou,
    }
    return render(request, 'core/project_detail.html', context)

# ---  Part 3/4: 顧客詳細 & マップ & 削除・完了 ---
@login_required
def assignment_detail(request, pk, assignment_pk):
    project = get_object_or_404(Project, pk=pk, allowed_users=request.user)
    assignment = get_object_or_404(Assignment, pk=assignment_pk, project=project)
    status_form = StatusForm(request.POST or None, instance=assignment)
    photo_form = PhotoForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and 'delete_photo' in request.POST:
        Photo.objects.filter(assignment=assignment, photo_type=request.POST['delete_photo']).delete()
        return redirect('core:assignment_detail', pk=pk, assignment_pk=assignment_pk)
    if request.method == 'POST' and request.FILES.get('image') and photo_form.is_valid():
        photo = photo_form.save(commit=False)
        photo.assignment = assignment
        photo.save()
        return redirect('core:assignment_detail', pk=pk, assignment_pk=assignment_pk)
    if request.method == 'POST' and status_form.is_valid():
        status_form.save()
        return redirect('core:assignment_detail', pk=pk, assignment_pk=assignment_pk)

    photo_slots = []
    for ptype, plabel in Photo.PHOTO_TYPE_CHOICES:
        existing = assignment.photos.filter(photo_type=ptype).first()
        photo_slots.append({'type': ptype, 'label': plabel, 'photo': existing})

    return render(request, 'core/assignment_detail.html', {
        'project': project,
        'assignment': assignment,
        'status_form': status_form,
        'photo_slots': photo_slots,
    })

@login_required
def project_map(request, pk):
    project = get_object_or_404(Project, pk=pk, allowed_users=request.user)
    assignments = Assignment.objects.filter(project=project).select_related('customer')
    floors = defaultdict(list)
    for a in assignments:
        room = a.customer.room_number or ''
        if room.isdigit(): floor = int(room[:-2]) if len(room)>=3 else int(room[0]) if len(room)==2 else 0
        else: floor = -1
        floors[floor].append(a)
    sorted_floors = sorted(floors.items(), key=lambda x: (x[0]==-1, -x[0] if x[0]!=-1 else 0))
    return render(request, 'core/project_map.html', {'project': project, 'grouped_assignments': sorted_floors})

@login_required
def project_delete(request, pk):
    proj = get_object_or_404(Project, pk=pk)
    if request.method == 'POST' and request.user.is_staff:
        proj.soft_delete()
        messages.success(request, '案件を削除フォルダに移動しました（30日後に自動削除）')
        return redirect('core:project_list')

@login_required
def project_complete(request, pk):
    proj = get_object_or_404(Project, pk=pk)
    if request.method == 'POST' and request.user.is_staff:
        proj.is_completed = True
        proj.save()
        messages.success(request, '案件を完了としてマークしました')
        return redirect('core:project_list')

# --- Part 4/4: ユーザー管理・編集・インポート・パスワード ---
@login_required
def user_manage(request):
    if not request.user.is_staff: return redirect('home')
    if request.method=='POST' and 'bulk_delete' in request.POST:
        selected = request.POST.getlist('selected_user')
        if not selected: messages.warning(request,'削除するユーザーを選択してください。')
        elif str(request.user.pk) in selected: messages.error(request,'自分自身は削除できません。')
        else: count,_ = User.objects.filter(pk__in=selected).delete(); messages.success(request,f'{count} 件削除')
        return redirect('core:user_manage')
    excel_form = ExcelUploadForm(request.POST or None, request.FILES or None)
    if request.method=='POST' and 'excel_file' in request.FILES and excel_form.is_valid():
        wb = load_workbook(request.FILES['excel_file'])
        sheet = wb.active
        headers = [c.value for c in sheet[1]]
        def idx(col):
            if col not in headers: raise ValueError(f"ヘッダー「{col}」が見つかりません")
            return headers.index(col)
        preview_data=[]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code       = str(row[idx('氏名コード')] or '').strip().zfill(7)
            last_name  = str(row[idx('姓')] or '').strip()
            first_name = str(row[idx('名')] or '').strip()
            company    = str(row[idx('会社')] or '').strip()
            district   = str(row[idx('地区')] or '').strip()
            team       = str(row[idx('チーム')] or '').strip()
            group      = str(row[idx('グループ')] or '').strip()
            is_staff   = str(row[idx('スタッフ権限')] or '').strip() == '1'
            exists     = User.objects.filter(code=code).exists()
            preview_data.append({'code':code,'last_name':last_name,'first_name':first_name,'company':company,'district':district,'team':team,'group':group,'is_staff':is_staff,'exists':exists})
        request.session['preview_data'] = preview_data
        return redirect('core:import_users_confirm')
    filter_form = UserFilterForm(request.GET or None)
    def mc(f): return [('', '全て')] + [(v, v) for v in User.objects.order_by(f).values_list(f, flat=True).distinct() if v]
    filter_form.fields['company'].choices  = mc('company')
    filter_form.fields['district'].choices = mc('district')
    filter_form.fields['team'].choices     = mc('team')
    filter_form.fields['group'].choices    = mc('group')
    users = User.objects.order_by('code')
    if filter_form.is_valid():
        cd = filter_form.cleaned_data
        for f in ['company','district','team','group']:
            if cd[f]: users = users.filter(**{f:cd[f]})
        if cd['is_staff'] in ['0','1']: users=users.filter(is_staff=(cd['is_staff']=='1'))
    return render(request,'core/user_manage.html',{'excel_form':excel_form,'filter_form':filter_form,'users':users})

@login_required
def user_edit(request, pk):
    if not request.user.is_staff: return redirect('home')
    usr = get_object_or_404(User, pk=pk)
    form = CustomUserForm(request.POST or None, instance=usr)
    if request.method=='POST' and form.is_valid(): form.save(); return redirect('core:user_manage')
    return render(request,'core/user_edit.html',{'form':form,'user_obj':usr})

@require_http_methods(["GET", "POST"])
def import_users_confirm(request):
    # セッションからプレビュー用データを取得
    preview_data = list(request.session.get('preview_data', []))

    # POST のときだけ処理を実行
    if request.method == 'POST':
        # ボタンアクションと行インデックスを取得
        action  = request.POST.get('action')
        idx_str = request.POST.get('index')

        # 編集ボタン → 編集画面へ
        if action == 'edit' and idx_str:
            request.session['edit_index'] = int(idx_str)
            return redirect('core:import_users_edit')

        # 削除ボタン → 該当行を削除して再表示
        if action == 'delete' and idx_str:
            i = int(idx_str)
            preview_data.pop(i)
            request.session['preview_data'] = preview_data
            messages.info(request, f"{i+1} 行目を削除しました。")
            return redirect('core:import_users_confirm')

        # 登録ボタン → ユーザーをインポート
        if action == 'register':
            imported, skipped = [], []
            for row in preview_data:
                code = row.get('code')
                # 既存ユーザーはスキップ
                if not code or User.objects.filter(Q(code=code) | Q(username=code)).exists():
                    skipped.append(code)
                    continue
                User.objects.create_user(
                    username=code,
                    password=code,
                    code=code,
                    last_name=row.get('last_name', ''),
                    first_name=row.get('first_name', ''),
                    company=row.get('company', ''),
                    district=row.get('district', ''),
                    team=row.get('team', ''),
                    group=row.get('group', ''),
                    is_staff=row.get('is_staff', False),
                )
                imported.append(code)
            # セッションをクリアしてメッセージ
            request.session.pop('preview_data', None)
            messages.success(
                request,
                f"{len(imported)} 件インポートしました。"
                + (f"{len(skipped)} 件スキップしました。" if skipped else "")
            )
            return redirect('core:user_manage')

    # GET またはその他のケースは確認画面を表示
    return render(request, 'core/import_users_confirm.html', {
        'preview_data': preview_data,
    })


@require_http_methods(["GET","POST"])
def import_users_edit(request):
    preview_data=list(request.session.get('preview_data',[]))
    idx=request.session.get('edit_index')
    if idx is None or not (0<=idx<len(preview_data)): messages.error(request,'対象行が存在しません'); return redirect('core:import_users_confirm')
    if request.method=='POST':
        form=UserImportEditForm(request.POST)
        if form.is_valid(): preview_data[idx]=form.cleaned_data; request.session['preview_data']=preview_data; request.session.pop('edit_index',None); messages.success(request,f"{idx+1} 行目更新"); return redirect('core:import_users_confirm')
    form=UserImportEditForm(initial=preview_data[idx])
    return render(request,'core/import_users_edit.html',{'form':form,'index':idx})

@login_required
def password_change(request):
    if request.method=='POST':
        form=PasswordChangeForm(request.user,request.POST)
        if form.is_valid(): user=form.save(); update_session_auth_hash(request,user); messages.success(request,'パスワードを変更しました'); return redirect('core:password_change_done')
    else:
        form=PasswordChangeForm(request.user)
    return render(request,'core/password_change.html',{'form':form})

@login_required
def password_change_done(request):
    return render(request,'core/password_change_done.html')

@login_required
def photo_upload(request, pk, assignment_pk):
    # 権限チェック付きで取得
    project    = get_object_or_404(Project, pk=pk, allowed_users=request.user)
    assignment = get_object_or_404(Assignment, pk=assignment_pk, project=project)

    if request.method == 'POST':
        form = PhotoForm(request.POST, request.FILES)
        if form.is_valid():
            photo = form.save(commit=False)
            photo.assignment = assignment
            photo.save()
            # アップロード後は詳細画面へ戻る
            return redirect('core:assignment_detail', pk=pk, assignment_pk=assignment_pk)
    else:
        form = PhotoForm()

    return render(request, 'core/photo_upload.html', {
        'form': form,
        'project': project,
        'assignment': assignment,
    })